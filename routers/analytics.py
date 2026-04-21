"""
Роутер для аналитики.
Все URL начинаются с /analytics
Обрабатывает:
- GET /analytics/summary - основные метрики
- GET /analytics/top-products - топ продуктов
- GET /analytics/summary-usd - метрики в долларах
- POST /analytics/upload-csv - загрузка продаж из CSV
"""

from fastapi import APIRouter, Query, HTTPException, UploadFile, File
from models.analytics import SummaryMetrics
from models.sale import Sale
from services.storage import add_sales, get_sales
from services.aggregation import calculate_summary, calculate_top_products
from services.currency import get_usd_rate
from typing import List
from datetime import date
import pandas as pd
from io import StringIO, BytesIO

# === для экселя ===========
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font

router = APIRouter(prefix="/analytics", tags=["analytics"])


@router.get("/summary", response_model=SummaryMetrics)
async def get_summary(
    # ...,Ellipsis (многоточие)	ОБЯЗАТЕЛЬНЫЙ параметр
    date_from: date = Query(
        ..., description="Начальная дата периода (YYYY-MM-DD)"
    ),
    date_to: date = Query(
        ..., description="Конечная дата периода (YYYY-MM-DD)"
    ),
    marketplace: str | None = Query(
        None,
        description="Фильтр по маркетплейсу (ozon, wildberries, yandex_market)",
    ),
) -> SummaryMetrics:
    """
    GET /analytics/summary
    Агрегированные метрики за период.

    date_from и date_to - обязательные параметры.
    marketplace - опциональный фильтр.

    Пример запроса:
    /analytics/summary?date_from=2025-03-01&date_to=2025-03-31&marketplace=ozon

    Пример ответа:
    {
        "total_revenue": 125000.50,
        "total_cost": 35000.25,
        "gross_profit": 90000.25,
        "margin_percent": 72.0,
        "total_orders": 42,
        "avg_order_value": 2976.20,
        "return_rate": 8.5
    }
    """

    # ========== ШАГ 1: Получаем данные из БД ==========
    # Здесь вызывается storage.get_sales()
    # Возвращает список объектов Sale
    # sales = [Sale(...), Sale(...), ...]
    sales = get_sales(
        marketplace=marketplace,
        date_from=date_from,
        date_to=date_to,
    )

    # ========== ШАГ 2: Считаем метрики и возвращаем метрики в json ==========
    # Здесь вызывается aggregation.calculate_summary()
    # Передаем список продаж, получаем метрики
    # metrics = SummaryMetrics(total_revenue=125000.50, ...)
    metrics = calculate_summary(sales)
    return metrics


@router.get("/top-products")
async def get_top_products(
    date_from: date = Query(
        ..., description="Начальная дата периода (YYYY-MM-DD)"
    ),
    date_to: date = Query(
        ..., description="Конечная дата периода (YYYY-MM-DD)"
    ),
    sort_by: str = Query(
        "revenue",
        # уууу. вот они и регулярки
        regex="^(revenue|quantity|profit)$",
        description="Критерий сортировки: revenue, quantity, profit",
    ),
    limit: int = Query(
        10, ge=1, le=50, description="Количество продуктов (макс 50)"
    ),
) -> List[dict]:
    """
    GET /analytics/top-products
    Топ продуктов за период.

    date_from и date_to - обязательные параметры.
    sort_by - по какому критерию сортировать.
    limit - сколько продуктов вернуть.

    Пример запроса:
    /analytics/top-products?date_from=2025-03-01&date_to=2025-03-31&sort_by=revenue&limit=5

    Ответ
    [
        {"product_name": "Кабель USB-C", "revenue": 12500.50, "quantity": 42, "profit": 8750.25},
        {"product_name": "Наушники TWS", "revenue": 10200.00, "quantity": 15, "profit": 7200.00},
        {"product_name": "Чехол для iPhone", "revenue": 8900.00, "quantity": 8, "profit": 6200.00},
        {"product_name": "Зарядка 65W", "revenue": 7500.00, "quantity": 10, "profit": 5100.00},
        {"product_name": "Подставка для ноутбука", "revenue": 6200.00, "quantity": 5, "profit": 4300.00}
    ]
    """

    # ========== ШАГ 1: Получаем данные из БД ==========
    sales = get_sales(
        date_from=date_from,
        date_to=date_to,
    )

    # ========== ШАГ 2: Считаем топ продуктов и возвращаем ==========
    # Здесь вызывается aggregation.calculate_top_products()
    # top_products = [{"product_name": "...", "revenue": ..., ...}, ...]
    top_products = calculate_top_products(sales, sort_by, limit)
    return top_products


@router.get("/summary-usd")
async def get_summary_usd(
    date_from: date = Query(
        ..., description="Начальная дата периода (YYYY-MM-DD)"
    ),
    date_to: date = Query(
        ..., description="Конечная дата периода (YYYY-MM-DD)"
    ),
    marketplace: str | None = Query(
        None, description="Фильтр по маркетплейсу"
    ),
) -> dict:
    """
    GET /analytics/summary-usd
    Агрегированные метрики за период в USD.

    Работает как /analytics/summary, но все денежные метрики конвертированы в USD.
    Курс берется из API ЦБ РФ и кэшируется на 1 час.

    Пример ответа:
    {
        "total_revenue": 1351.35,
        "total_cost": 378.38,
        "gross_profit": 972.97,
        "margin_percent": 72.0,
        "total_orders": 42,
        "avg_order_value": 32.17,
        "return_rate": 8.5,
        "usd_rate": 92.50
    }
    """

    # ========== ШАГ 1: Получаем курс USD ==========
    try:
        usd_rate = await get_usd_rate()
    except Exception:
        raise HTTPException(status_code=503, detail="Currency API unavailable")

    # ========== ШАГ 2: Получаем метрики в рублях ==========
    # Переиспользуем функцию get_summary (которая уже делает всё нужное)
    rub_metrics = await get_summary(date_from, date_to, marketplace)

    # ========== ШАГ 3: Конвертируем в USD ==========
    return {
        "total_revenue": round(rub_metrics.total_revenue / usd_rate, 2),
        "total_cost": round(rub_metrics.total_cost / usd_rate, 2),
        "gross_profit": round(rub_metrics.gross_profit / usd_rate, 2),
        "margin_percent": rub_metrics.margin_percent,  # Процент не меняется
        "total_orders": rub_metrics.total_orders,  # Количество не меняется
        "avg_order_value": round(rub_metrics.avg_order_value / usd_rate, 2),
        "return_rate": rub_metrics.return_rate,  # Процент не меняется
        "usd_rate": round(usd_rate, 2),  # Добавляем курс в ответ
    }


@router.post("/upload-csv")
async def upload_csv(
    file: UploadFile = File(..., description="CSV файл с данными о продажах")
) -> dict:
    """
    POST /analytics/upload-csv
    Загрузка продаж из CSV-файла.

    Принимает CSV файл со столбцами:
    - order_id: номер заказа
    - marketplace: ozon, wildberries, yandex_market
    - product_name: название продукта
    - quantity: количество
    - price: цена за штуку
    - cost_price: себестоимость за штуку
    - status: delivered, returned, cancelled
    - sold_at: дата продажи (YYYY-MM-DD)

    Returns:
        - loaded_count: количество успешно загруженных строк
        - error_count: количество строк с ошибками
        - errors: список ошибок с номерами строк

    Example:
        curl -X POST "http://localhost:8000/analytics/upload-csv" \
             -F "file=@sample_data.csv"

    Example response:
        {
            "loaded_count": 18,
            "error_count": 2,
            "errors": [
                {"row": 5, "error": "price must be > 0"},
                {"row": 12, "error": "sold_at must be a valid date"}
            ]
        }
    """

    # ========== ШАГ 1: Проверяем формат файла ==========
    if not file.filename.endswith('.csv'):
        raise HTTPException(
            status_code=400, detail="Only CSV files are allowed"
        )

    # ========== ШАГ 2: Читаем CSV файл ==========
    try:
        content = await file.read()
        # Декодируем байты в строку и читаем через pandas
        df = pd.read_csv(StringIO(content.decode('utf-8')))
    except Exception as e:
        raise HTTPException(
            status_code=400, detail=f"Error reading CSV file: {str(e)}"
        )

    # ========== ШАГ 3: Проверяем наличие обязательных колонок ==========
    required_columns = [
        'order_id',
        'marketplace',
        'product_name',
        'quantity',
        'price',
        'cost_price',
        'status',
        'sold_at',
    ]

    missing_columns = [
        col for col in required_columns if col not in df.columns
    ]
    if missing_columns:
        raise HTTPException(
            status_code=400,
            detail=f"Missing columns: {', '.join(missing_columns)}",
        )

    # ========== ШАГ 4: Валидируем каждую строку ==========
    valid_sales = []
    errors = []

    for idx, row in df.iterrows():
        try:
            # Пробуем создать объект Sale (Pydantic сам провалидирует)
            sale = Sale(
                order_id=str(row['order_id']),
                marketplace=row['marketplace'],
                product_name=row['product_name'],
                quantity=int(row['quantity']),
                price=float(row['price']),
                cost_price=float(row['cost_price']),
                status=row['status'],
                sold_at=date.fromisoformat(
                    row['sold_at']
                ),  # преобразуем строку в date
            )
            valid_sales.append(sale)
        except Exception as e:
            # Номер строки +2 потому что:
            # - idx начинается с 0
            # - первая строка (индекс 0) - это заголовки
            # - первая строка данных имеет номер 2 в CSV
            errors.append({"row": idx + 2, "error": str(e)})

    # ========== ШАГ 5: Сохраняем валидные продажи ==========
    loaded_count = 0
    if valid_sales:
        loaded_count = add_sales(valid_sales)

    # ========== ШАГ 6: Возвращаем результат ==========
    return {
        "loaded_count": loaded_count,
        "error_count": len(errors),
        "errors": errors[:20],  # возвращаем не более 20 ошибок
    }


# ========== НОВЫЕ ЭНДПОИНТЫ ДЛЯ EXCEL ==========


# ========== НОВЫЕ ЭНДПОИНТЫ ДЛЯ EXCEL ==========


@router.get("/export-raw-data")
async def export_raw_data_to_excel() -> StreamingResponse:
    """
    GET /analytics/export-raw-data
    Выгружает все продажи из базы в Excel-файл
    """

    # 1. Получаем все продажи из БД
    sales = get_sales()  # без фильтров, все данные

    # 2. Создаем Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Продажи"

    # 3. Записываем заголовки
    ws.append(
        [
            "order_id",
            "marketplace",
            "product_name",
            "quantity",
            "price",
            "cost_price",
            "status",
            "sold_at",
        ]
    )

    # 4. Записываем данные
    for sale in sales:
        ws.append(
            [
                sale.order_id,
                sale.marketplace,
                sale.product_name,
                sale.quantity,
                sale.price,
                sale.cost_price,
                sale.status,
                sale.sold_at.isoformat(),
            ]
        )

    # 5. Авто-подбор ширины колонок
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    # 6. Сохраняем в память и отдаем
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=sales_data.xlsx"
        },
    )


@router.get("/export-analytics")
async def export_analytics_to_excel(
    date_from: date = Query(..., description="Начальная дата"),
    date_to: date = Query(..., description="Конечная дата"),
    marketplace: str | None = Query(None),
) -> StreamingResponse:
    """
    GET /analytics/export-analytics
    Выгружает аналитические метрики в Excel-файл
    """

    # 1. Получаем данные
    sales = get_sales(
        marketplace=marketplace, date_from=date_from, date_to=date_to
    )
    metrics = calculate_summary(sales)
    top_products = calculate_top_products(sales, limit=10)

    # 2. Создаем Excel-файл
    wb = Workbook()

    # Лист 1: Сводные метрики
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    ws_summary.append(["Метрика", "Значение"])
    ws_summary.append(["total_revenue", metrics.total_revenue])
    ws_summary.append(["total_cost", metrics.total_cost])
    ws_summary.append(["gross_profit", metrics.gross_profit])
    ws_summary.append(["margin_percent", f"{metrics.margin_percent}%"])
    ws_summary.append(["total_orders", metrics.total_orders])
    ws_summary.append(["avg_order_value", metrics.avg_order_value])
    ws_summary.append(["return_rate", f"{metrics.return_rate}%"])

    # Форматирование сводки
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 20
    ws_summary["A1"].font = Font(bold=True)
    ws_summary["B1"].font = Font(bold=True)

    # Лист 2: Топ продуктов
    ws_products = wb.create_sheet("Топ продуктов")
    ws_products.append(["Продукт", "Выручка", "Количество", "Прибыль"])

    for p in top_products:
        ws_products.append(
            [p["product_name"], p["revenue"], p["quantity"], p["profit"]]
        )

    # Форматирование таблицы
    ws_products.column_dimensions["A"].width = 25
    ws_products.column_dimensions["B"].width = 15
    ws_products.column_dimensions["C"].width = 15
    ws_products.column_dimensions["D"].width = 15
    ws_products["A1"].font = Font(bold=True)
    ws_products["B1"].font = Font(bold=True)
    ws_products["C1"].font = Font(bold=True)
    ws_products["D1"].font = Font(bold=True)

    # 3. Сохраняем и отдаем
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=analytics_{date_from}_to_{date_to}.xlsx"
        },
    )


@router.get("/export-analytics-with-macro")
async def export_analytics_with_macro_instruction(
    date_from: date = Query(..., description="Начальная дата"),
    date_to: date = Query(..., description="Конечная дата"),
    marketplace: str | None = Query(None),
) -> StreamingResponse:
    """
    GET /analytics/export-analytics-with-macro
    Выгружает аналитику в Excel-файл с листом-инструкцией и VBA-кодом
    """

    # 1. Получаем данные
    sales = get_sales(
        marketplace=marketplace, date_from=date_from, date_to=date_to
    )
    metrics = calculate_summary(sales)
    top_products = calculate_top_products(sales, limit=10)

    # 2. Создаем Excel-файл
    wb = Workbook()

    # ========== ЛИСТ 1: Сводка ==========
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    ws_summary.append(["Метрика", "Значение"])
    ws_summary.append(["total_revenue", metrics.total_revenue])
    ws_summary.append(["total_cost", metrics.total_cost])
    ws_summary.append(["gross_profit", metrics.gross_profit])
    ws_summary.append(["margin_percent", f"{metrics.margin_percent}%"])
    ws_summary.append(["total_orders", metrics.total_orders])
    ws_summary.append(["avg_order_value", metrics.avg_order_value])
    ws_summary.append(["return_rate", f"{metrics.return_rate}%"])

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 20
    ws_summary["A1"].font = Font(bold=True)
    ws_summary["B1"].font = Font(bold=True)

    # ========== ЛИСТ 2: Топ продуктов ==========
    ws_products = wb.create_sheet("Топ продуктов")
    ws_products.append(["Продукт", "Выручка", "Количество", "Прибыль"])

    for p in top_products:
        ws_products.append(
            [p["product_name"], p["revenue"], p["quantity"], p["profit"]]
        )

    ws_products.column_dimensions["A"].width = 25
    ws_products.column_dimensions["B"].width = 15
    ws_products.column_dimensions["C"].width = 15
    ws_products.column_dimensions["D"].width = 15
    ws_products["A1"].font = Font(bold=True)
    ws_products["B1"].font = Font(bold=True)
    ws_products["C1"].font = Font(bold=True)
    ws_products["D1"].font = Font(bold=True)

    # ========== ЛИСТ 3: Инструкция + VBA-код (с подробными комментариями) ==========
    ws_instruction = wb.create_sheet("Инструкция")

    ws_instruction["A1"] = (
        "ИНСТРУКЦИЯ ПО АВТОМАТИЧЕСКОМУ ФОРМАТИРОВАНИЮ ОТЧЕТА"
    )
    ws_instruction["A1"].font = Font(bold=True, size=14)
    ws_instruction.merge_cells("A1:E1")

    ws_instruction["A3"] = "ШАГ 1: Открыть редактор VBA"
    ws_instruction["A3"].font = Font(bold=True)
    ws_instruction["A4"] = (
        "Нажмите Alt + F11 (Windows) или Fn + Option + F11 (Mac)"
    )

    ws_instruction["A6"] = "ШАГ 2: Вставить новый модуль"
    ws_instruction["A6"].font = Font(bold=True)
    ws_instruction["A7"] = 'В меню редактора выберите: Insert → Module'

    ws_instruction["A9"] = "ШАГ 3: Скопировать код макроса"
    ws_instruction["A9"].font = Font(bold=True)
    ws_instruction["A10"] = (
        "Скопируйте код из ячеек ниже и вставьте его в открывшееся окно модуля"
    )

    ws_instruction["A32"] = "ШАГ 4: Запустить макрос"
    ws_instruction["A32"].font = Font(bold=True)
    ws_instruction["A33"] = (
        'Закройте редактор VBA. В Excel нажмите Alt + F8, выберите макрос "FormatReport" и нажмите "Выполнить"'
    )

    # VBA-код макроса с ПОДРОБНЫМИ КОММЕНТАРИЯМИ на русском
    vba_code_lines = [
        "' =============================================================",
        "' НАЧАЛО МАКРОСА",
        "' =============================================================",
        "'",
        "' Этот макрос автоматически форматирует отчет:",
        "'   - Подбирает ширину колонок",
        "'   - Делает заголовки жирными",
        "'   - Заливает заголовки зеленым цветом",
        "'   - Добавляет границы к таблице продуктов",
        "'   - Показывает сообщение об успешном завершении",
        "'",
        "' Как установить:",
        "'   1. Нажмите Alt+F11 (откроется редактор VBA)",
        "'   2. В меню: Insert → Module",
        "'   3. Скопируйте этот код в открывшееся окно",
        "'   4. Закройте редактор",
        "'   5. В Excel нажмите Alt+F8 → FormatReport → Выполнить",
        "'",
        "' =============================================================",
        "",
        "Sub FormatReport()",
        "    ' ---------------------------------------------------------",
        "    ' ОБЪЯВЛЕНИЕ ПЕРЕМЕННЫХ",
        "    ' ---------------------------------------------------------",
        "    Dim ws1 As Worksheet   ' переменная для листа \"Сводка\"",
        "    Dim ws2 As Worksheet   ' переменная для листа \"Топ продуктов\"",
        "    Dim lastRow As Long     ' переменная для хранения номера последней строки",
        "    ",
        "    ' ---------------------------------------------------------",
        "    ' ПРИВЯЗКА К ЛИСТАМ (присваиваем переменным реальные листы)",
        "    ' ---------------------------------------------------------",
        "    Set ws1 = ThisWorkbook.Sheets(\"Сводка\")",
        "    Set ws2 = ThisWorkbook.Sheets(\"Топ продуктов\")",
        "    ",
        "    ' ---------------------------------------------------------",
        "    ' ФОРМАТИРОВАНИЕ ЛИСТА \"Сводка\"",
        "    ' ---------------------------------------------------------",
        "    ' AutoFit - автоматически подбирает ширину колонок",
        "    ws1.Columns(\"A:B\").AutoFit",
        "    ",
        "    ' Bold = True - делает текст жирным",
        "    ws1.Rows(1).Font.Bold = True",
        "    ",
        "    ' Interior.Color - заливает ячейку цветом",
        "    ' RGB(144, 238, 144) - светло-зеленый (light green)",
        "    ws1.Rows(1).Interior.Color = RGB(144, 238, 144)",
        "    ",
        "    ' ---------------------------------------------------------",
        "    ' ФОРМАТИРОВАНИЕ ЛИСТА \"Топ продуктов\"",
        "    ' ---------------------------------------------------------",
        "    ws2.Columns(\"A:D\").AutoFit",
        "    ws2.Rows(1).Font.Bold = True",
        "    ws2.Rows(1).Interior.Color = RGB(144, 238, 144)",
        "    ",
        "    ' ---------------------------------------------------------",
        "    ' ДОБАВЛЕНИЕ ГРАНИЦ К ТАБЛИЦЕ ПРОДУКТОВ",
        "    ' ---------------------------------------------------------",
        "    ' Cells(ws2.Rows.Count, 1) - последняя ячейка в колонке A",
        "    ' End(xlUp) - поднимается вверх до последней заполненной ячейки",
        "    ' Row - возвращает номер строки",
        "    lastRow = ws2.Cells(ws2.Rows.Count, 1).End(xlUp).Row",
        "    ",
        "    ' Если таблица не пустая (есть хотя бы строка заголовка + данные)",
        "    If lastRow > 1 Then",
        "        ' Range(\"A1:D\" & lastRow) - диапазон от A1 до последней строки колонки D",
        "        ' Borders.LineStyle = xlContinuous - сплошная линия границы",
        "        ' Borders.Weight = xlThin - тонкая линия",
        "        With ws2.Range(\"A1:D\" & lastRow)",
        "            .Borders.LineStyle = xlContinuous",
        "            .Borders.Weight = xlThin",
        "        End With",
        "    End If",
        "    ",
        "    ' ---------------------------------------------------------",
        "    ' СООБЩЕНИЕ ОБ УСПЕШНОМ ЗАВЕРШЕНИИ",
        "    ' ---------------------------------------------------------",
        "    ' MsgBox - показывает всплывающее окно",
        "    ' vbInformation - иконка \"i\" (информация)",
        "    MsgBox \"Отчет успешно отформатирован!\", vbInformation, \"Аналитика\"",
        "End Sub",
        "",
        "Sub Auto_Open()",
        "    ' =============================================================",
        "    ' ЭТА ФУНКЦИЯ ЗАПУСКАЕТСЯ АВТОМАТИЧЕСКИ ПРИ ОТКРЫТИИ ФАЙЛА",
        "    ' =============================================================",
        "    ' Вызывает главный макрос форматирования",
        "    Call FormatReport",
        "End Sub",
        "",
        "' =============================================================",
        "' КОНЕЦ МАКРОСА",
        "' =============================================================",
    ]

    for i, line in enumerate(vba_code_lines):
        ws_instruction.cell(row=13 + i, column=2, value=line)

    ws_instruction.column_dimensions["A"].width = 40
    ws_instruction.column_dimensions["B"].width = 70

    # 3. Сохраняем и отдаем
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=analytics_{date_from}_to_{date_to}.xlsx"
        },
    )
